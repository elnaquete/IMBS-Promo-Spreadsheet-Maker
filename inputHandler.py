from openpyxl import load_workbook

def strToBool (str):
    '''
    IN: String: true, TRUE, True, tRUE or false, FALSE, False, fALSE
    OUT: Booleans for that strings
    '''
    #Ojo: para pasar los 'True' a booleanos, hay que usar un if, no hay un casting directo. (cualquier string
    #con contenido evalua a True)
    if str == None:
        return False
    elif str.lower() == 'no' or str == '':
        return False
    else:
        return True


def readExcel (file = 'input.xlsx'):
    '''
    este archivo se encarga de leer el excel con las promos cargadas, rechazar las incorrectas.x ej:
    -EndDate menor que StartDate
    -FEED no se encuentra en lista de feeds
    -Idem con Promo Pckg
    (Ver si no podemos hacer esto desde la validacion de Excel)

    OUTPUT: lista de diccionarios, cada diccionario es una promo.
      

    POSIBLES PARAMETROS A AGREGAR: 
    Acordarse de las promos solo digitales que no van al aire (ej. TempDeJuegos - showFeed = 'OFFAIR'

    '''
    wb = load_workbook(filename=file, read_only=True)
    datos = wb['Data'] 

    promo = {} #diccionario, cada promo es uno
    output = [] #lista con todas las promos ingresadas

    #itero todas las filas del Excel, creando un diccionario por cada fila. 
    #Voy llenando la lista 'output' con cada diccionario. 

    #agregar flag bumps? Algo del tipo "NEcesita bumps" (bool) o algo así?
    
    for row in datos.iter_rows (min_row=2, values_only=True): 
        if row[0] != None: #si la fila no esta vacia
            #hago un if para que si es "AMC*4" me multiplique las promos *4?
            if row[0] == 'AMC*4FEEDS':
                amcFeeds = ('AMCBRASIL', 'AMCLATAM', 'AMCNORCOL', 'AMCSUR')
                for feed in amcFeeds:
                    promo = {
                    'showFeed': feed,
                    'showName': row[1], 
                    'promoPckg': row[2], 
                    'duration': 30 if row[3] == None else int(row[3]),
                    'premiereDate': row[4],
                    'genDateStr': row[5], 
                    'genStartDate': row[6],
                    'endDate': row[7],    
                    'dstMex': strToBool(row[8]), 
                    'dstChi':strToBool(row[9]), 
                    'crossChannel':strToBool(row[10]),
                    'megaCable': strToBool(row[11]),
                    'a&e': strToBool(row[12]),
                    'cines': strToBool(row[13]),
                    'foxSports': strToBool(row[14])
                    }
                    output.append(promo)
            else:
                promo = {
                'showFeed': row[0],
                'showName': row[1], 
                'promoPckg': row[2],
                'duration': 30 if row[3] == None else int(row[3]),
                'premiereDate': row[4],
                'genDateStr': row[5], 
                'genStartDate': row[6],
                'endDate': row[7],    
                'dstMex': strToBool(row[8]), 
                'dstChi':strToBool(row[9]), 
                'crossChannel':strToBool(row[10]),
                'megaCable': strToBool(row[11]),
                'a&e': strToBool(row[12]),
                'cines': strToBool(row[13]),
                'foxSports': strToBool(row[14])
                }
                output.append(promo)
    
    return output

#print(readExcel())
 