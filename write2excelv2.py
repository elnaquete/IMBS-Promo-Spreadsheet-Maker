def columnTidier(workbook):  # Puesto a mano, luego vemos de automatizarlo.
    for sheet in workbook:
        sheet.column_dimensions["A"].width = 50
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 50
        sheet.column_dimensions["F"].width = 17
        sheet.column_dimensions["H"].width = 3
        sheet.column_dimensions["I"].width = 10
        sheet.column_dimensions["J"].width = 10
        sheet.column_dimensions["K"].width = 10
        sheet.column_dimensions["L"].width = 15


def write2excelIBMSv2(listaPromos, dest_filename):
    """
    IN: promoList = (list) list including each promo (dict)
    filename = (str) excel filename (include .xlsx extension)
    OUT: Writes Excel file, with separate sheet for each feed
    """
    encabezadoIBMS = [
        "Promo Name",
        "Time Code In",
        "Time Code Out",
        "Length",
        "Detail",
        "Feed",
        "MainMI",
        "AFP",
        "START",
        "END",
        "DUE DATE",
        "Allowed Days",
    ]
    from openpyxl import Workbook
    from listMaker import IBMSlistMaker

    wb = Workbook()  # creo libro nuevo .
    # la plani por default se llama 'Sheet'. Al final de todo, hay que borrarla, porque esta vacía

    for promo in listaPromos:
        showFeed = promo["showFeed"]
        if promo["promoPckg"] == "BUMP":
            promoPack = IBMSlistMaker(promo)
            showFeed = "BUMPS " + showFeed
            for promoRow in promoPack:
                if "Disculpe" in promoRow[0]:  # Esto no anda, arreglarlo
                    print(promoRow[0])
                    if "ALERTAS" not in wb.sheetnames:
                        wb.create_sheet(title="ALERTAS")  # creo hoja nueva de alertas
                        ws1 = wb["ALERTAS"]
                        ws1.append(promoRow)  # la agrego como fila a la hoja de alertas
                    else:
                        ws1 = wb["ALERTAS"]
                        ws1.append(promoRow)  # la agrego como fila a la hoja de alertas
                else:
                    if (
                        showFeed not in wb.sheetnames
                    ):  # si no hay una hoja con el nombre de esa señal
                        wb.create_sheet(
                            title=showFeed
                        )  # creo hoja nueva con el nombre de esa señal
                        ws1 = wb[
                            showFeed
                        ]  # asigno a la variable ws1 el nombre de esa plani
                        ws1.append(
                            encabezadoIBMS
                        )  # agrego el encabezado al comienzo de la lista
                        ws1.append(promoRow)  # la agrego como fila a la planilla ws1
                    else:
                        ws1 = wb[
                            showFeed
                        ]  # asigno a la variable ws1 el nombre de esa plani
                        ws1.append(promoRow)  # la agrego como fila a la planilla ws1
        # aca podemos detectar el flag Crosschannel y armar la plani de Cross ;)
        else:
            promoPack = IBMSlistMaker(promo)
            for promoRow in promoPack:
                if "Disculpe" in promoRow[0]:  # Esto no anda, arreglarlo
                    print(promoRow[0])
                    if "ALERTAS" not in wb.sheetnames:
                        wb.create_sheet(title="ALERTAS")  # creo hoja nueva de alertas
                        ws1 = wb["ALERTAS"]
                        ws1.append(promoRow)  # la agrego como fila a la hoja de alertas
                    else:
                        ws1 = wb["ALERTAS"]
                        ws1.append(promoRow)  # la agrego como fila a la hoja de alertas
                else:
                    if (
                        showFeed not in wb.sheetnames
                    ):  # si no hay una hoja con el nombre de esa señal
                        wb.create_sheet(
                            title=showFeed
                        )  # creo hoja nueva con el nombre de esa señal
                        ws1 = wb[
                            showFeed
                        ]  # asigno a la variable ws1 el nombre de esa plani
                        ws1.append(
                            encabezadoIBMS
                        )  # agrego el encabezado al comienzo de la lista
                        ws1.append(promoRow)  # la agrego como fila a la planilla ws1
                    else:
                        ws1 = wb[
                            showFeed
                        ]  # asigno a la variable ws1 el nombre de esa plani
                        ws1.append(promoRow)  # la agrego como fila a la planilla ws1

    # Al final de todo, borrar la Hoja 'Sheet' que esta vacía
    wb.remove(wb["Sheet"])

    # Ajustar (manualmente) el ancho de las columnas de todas las hojas
    columnTidier(wb)
    # creo el archivo ("w" especifica que si no existe, lo cree)
    f = open(dest_filename, "w+")
    # Salvo el libro
    wb.save(filename=dest_filename)
