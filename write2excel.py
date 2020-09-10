def write2excelIBMS (list2print, dest_filename):
    '''
    IN: list = List of lists
    filename = (str) excel filename (include .xlsx extension)
    OUT: Writes Excel file
    '''
    encabezadoIBMS = [
    'Promo Name', 'Time Code In', 'Time Code Out', 'Length', 'Detail',
    'Feed', 'MainMI', 'AFP', 'START', 'END', 'DUE DATE', 'Allowed Days'
    ]
    from openpyxl import Workbook

    wb = Workbook() #creo libro nuevo . 
    #la plani por default se llama 'Sheet'. Al final de todo, hay que borrarla, porque esta vacía

    for item in list2print: #itero la lista de listas (PROMOS)
        if ''.join(item[0]) not in wb.sheetnames: #si no hay una hoja con el nombre de esa señal
            if item[0] == ['ALERTAS']: #si es un mensaje de alerta
                sheetTitle = ''.join(item[0]) #le doy a la hoja el nombre del primer elemento (header)
                wb.create_sheet(title = sheetTitle)  #creo hoja nueva con el nombre de esa señal
                ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
            else:
                sheetTitle = ''.join(item[0]) 
                wb.create_sheet(title = sheetTitle)  #creo hoja nueva con el nombre de esa señal
                wb[sheetTitle].append(encabezadoIBMS) #agrego el encabezado al comienzo de la lista
                ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
        sheetTitle = ''.join(item[0])
        ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
        for promo in item[1:]: #itero promo por promo
            ws1.append(promo)  #la agrego como fila a la planilla ws1
        #aca habria que arreglar el ancho de las columnas

    # Al final de todo, borrar la Hoja 'Sheet' que esta vacía
    wb.remove(wb['Sheet'])

    # Ajustar el ancho de las columnas de todas las hojas // NO FUNCIONA, VER POR QUE
    columnTidier(wb)
    #Salvo el libro
    wb.save(filename = dest_filename)






def write2excelSeguimiento (list2print, dest_filename):
    '''
    IN: list = List of lists
    filename = (str) excel filename (include .xlsx extension)
    OUT: Writes Excel file
    '''
    #Definir el encabezado - PENDIENTE
    encabezado = [
    'Promo Name', 
    ]
    from openpyxl import Workbook

    wb = Workbook() #creo libro nuevo . 
    #la plani por default se llama 'Sheet'. Al final de todo, hay que borrarla, porque esta vacía

    for item in list2print: #itero la lista de listas (PROMOS)
        if ''.join(item[0]) not in wb.sheetnames: #si no hay una hoja con el nombre de esa señal
            if item[0] == ['ALERTAS']: #si es un mensaje de alerta
                sheetTitle = ''.join(item[0]) #le doy a la hoja el nombre del primer elemento (header)
                wb.create_sheet(title = sheetTitle)  #creo hoja nueva con el nombre de esa señal
                ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
            else:
                sheetTitle = ''.join(item[0])
                wb.create_sheet(title = sheetTitle)  #creo hoja nueva con el nombre de esa señal
                wb[sheetTitle].append(encabezado) #agrego el encabezado al comienzo de la lista
                ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
        else:
            sheetTitle = ''.join(item[0])
            ws1 = wb[sheetTitle] #asigno a la variable ws1 el nombre de esa plani
            for promo in item[1:]: #itero promo por promo
                ws1.append(promo)  #la agrego como fila a la planilla ws1
    # Al final de todo, borrar la Hoja 'Sheet' que esta vacía
    wb.remove(wb['Sheet'])


def columnTidier (workbook):  #Puesto a mano, luego vemos de automatizarlo.
    for sheet in workbook:
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['E'].width = 40
        sheet.column_dimensions['F'].width = 17
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 15

    #Esta es la version que no pude hacer andar
    # from openpyxl.utils import get_column_letter
  

    # column_widths = []
    # for row in worksheet:
    #     for i, cell in enumerate(row):
    #         if len(column_widths) > i:
    #             if len(cell) > column_widths[i]:
    #                 column_widths[i] = len(cell)
    #         else:
    #             column_widths += [len(cell)]

    # for i, column_width in enumerate(column_widths):
    #     worksheet.column_dimensions[get_column_letter(i+1)].width = column_width

