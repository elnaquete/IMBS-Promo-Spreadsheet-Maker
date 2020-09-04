#para abrir libro ya existente
from openpyxl import load_workbook

def print_rows(sheet):
    
    for row in sheet.iter_rows(values_only=True):
        print(row)

def print_workbook (wb2):
    '''
    in: Excel workbook
    out: tuples. Iterates de whole Excel workbook, prints out all sheet and values
    '''
    for sheet in wb2.sheetnames:
      print (sheet)
      print (print_rows(wb2[sheet]))

wb2 = load_workbook('prueba.xlsx')

print (wb2.sheetnames) #imprime todos los nombres de las hojas
operations_sheet = wb2.create_sheet("Operations")
print (wb2.sheetnames) #imprime todos los nombres de las hojas
wb2.remove(operations_sheet) #borro la hoja "Operations"
print (wb2.sheetnames) #imprime todos los nombres de las hojas
#wb2.remove(wb2[Sheet]) #borro la hoja "Sheet" OJO! Hasta que no escribo los datos, no se guarda!
#print (wb2.sheetnames) #imprime todos los nombres de las hojas




# sheets = [] #lista de hojas de la plani
# #aca itero Libro / Filas / Valores en cada fila
# for sheet in wb:
#   sheets.append(sheet.title)  #armo una lista con las hojas activas



#print (sheets)



#for row in sheet.values:
#   for value in row:
#     print(value)



# for row in wb2.values:
#   for value in row:
#     print(value)
