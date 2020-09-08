
def columnTidier (worksheet):  #NO FUNCIONA, VER POR QUÉ
    from openpyxl.utils import get_column_letter

    column_widths = []
    for row in worksheet:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = column_width

from openpyxl import load_workbook

wb = load_workbook("lista_IBMS.xlsx")
for sheet in wb:
    print (sheet.title)
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['E'].width = 40
    sheet.column_dimensions['F'].width = 17
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 10
    sheet.column_dimensions['K'].width = 10
#    columnTidier 

wb.save(filename = "lista_IBMS.xlsx")


    